import streamlit as st
import pandas as pd
from io import StringIO, BytesIO
from pathlib import Path
import csv
import warnings
import tempfile
import os
from processor import process_workbook, load_codes_map

st.set_page_config(page_title="DEBUG - Controlli Fine Mese", layout="wide")
st.title("DEBUG: Controlli Fine Mese - diagnostica lettura file")

# --- utility per parsing robusto CSV/TXT ---
def detect_separator_by_count(text, n_lines=30):
    lines = text.splitlines()[:n_lines]
    counts = {sep: 0 for sep in ["\t", ";", ",", "|"]}
    for L in lines:
        for sep in counts:
            counts[sep] += L.count(sep)
    best = max(counts.items(), key=lambda kv: kv[1])
    return best[0] if best[1] > 0 else None

def try_parse_csv_text(text, warnings_list):
    # proviamo più strategie di parsing e torniamo (df, format_str)
    # 1) detect by count
    best_sep = detect_separator_by_count(text, n_lines=50)
    if best_sep:
        try:
            df = pd.read_csv(StringIO(text), sep=best_sep, engine="python", header=None, dtype=str, on_bad_lines="warn")
            return df, f"csv_detected_by_count_sep_{repr(best_sep)}"
        except Exception as e:
            warnings_list.append(f"read_csv with sep={best_sep!r} failed: {e}")

    # 2) try common separators in order
    for sep in ["\t", ";", ",", "|"]:
        try:
            df = pd.read_csv(StringIO(text), sep=sep, engine="python", header=None, dtype=str, on_bad_lines="warn")
            return df, f"csv_try_sep_{repr(sep)}"
        except Exception as e:
            warnings_list.append(f"read_csv sep={sep!r} failed: {e}")

    # 3) whitespace regex
    try:
        df = pd.read_csv(StringIO(text), sep=r'\s+', engine="python", header=None, dtype=str, on_bad_lines="warn")
        return df, "csv_sep_regex_whitespace"
    except Exception as e:
        warnings_list.append(f"read_csv sep=r'\\s+' failed: {e}")

    # 4) manual fallback riga-per-riga
    try:
        import csv as _csv
        lines = text.splitlines()
        rows = []
        maxcols = 0
        for i, line in enumerate(lines, start=1):
            parsed = None
            for sep in ["\t", ";", ",", "|"]:
                try:
                    parsed = next(_csv.reader([line], delimiter=sep))
                    if len(parsed) >= 1:
                        break
                except Exception:
                    parsed = None
            if parsed is None:
                parsed = [line]
            rows.append(parsed)
            if len(parsed) > maxcols:
                maxcols = len(parsed)
        norm_rows = [r + [""] * (maxcols - len(r)) for r in rows]
        df = pd.DataFrame(norm_rows).astype(str)
        return df, "csv_manual_fallback"
    except Exception as e:
        warnings_list.append(f"manual csv fallback failed: {e}")
        raise

def find_header_row_candidate(df, keywords=None, max_scan=8):
    if keywords is None:
        keywords = ["Residenza", "Matricola", "Cognome", "Nome", "Gruppo", "Data", "Turno", "TurnoC", "TurnoE"]
    for i in range(min(max_scan, df.shape[0])):
        row = df.iloc[i].astype(str).tolist()
        hit_count = sum(1 for k in keywords if any(k in (str(cell)) for cell in row))
        if hit_count >= 2:
            return i
    return None

# -----------------------------------------------------
uploaded_file = st.file_uploader(
    "Carica il file (.xls/.xlsx/.csv/.txt)",
    type=["xls", "xlsx", "csv", "txt"],
    accept_multiple_files=False
)
if uploaded_file is None:
    st.info("Carica un file per iniziare.")
    st.stop()

# Leggiamo i bytes
try:
    raw_bytes = uploaded_file.read()
except Exception:
    try:
        raw_bytes = uploaded_file.getvalue()
    except Exception as e:
        st.error(f"Impossibile leggere il file caricato: {e}")
        st.stop()

st.write("Lunghezza bytes:", len(raw_bytes))

# mostra prime bytes (hex) per riconoscere formato
st.write("Prime 64 bytes (hex):", raw_bytes[:64].hex())

# semplice riconoscimento "magic"
magic = raw_bytes[:8]
is_xlsx_like = False
if magic.startswith(b'PK'):
    st.write("Probabile ZIP container (xlsx o docx/pptx) -> formato XLSX probabile (PK..).")
    is_xlsx_like = True
elif magic.startswith(b'\xd0\xcf\x11\xe0'):
    st.write("Probabile file BIFF (vecchio .xls).")
    is_xlsx_like = True
else:
    try:
        raw_bytes.decode("utf-8")
        st.write("Sembra testo UTF-8 (potrebbe essere CSV/TXT).")
    except Exception:
        st.write("Sembra testo non-UTF8; proveremo diverse decodifiche/encodings.")

# salvo temporaneamente il file su disco per testare come file path
suffix = ".bin"
if raw_bytes[:2] == b'PK':
    suffix = ".xlsx"
elif raw_bytes[:4] == b'\xd0\xcf\x11\xe0':
    suffix = ".xls"
else:
    suffix = ".csv"

tmp_dir = Path(tempfile.gettempdir())
tmp_path = tmp_dir / f"tmp_upload_debug{suffix}"
tmp_path.write_bytes(raw_bytes)
st.write("Saved tmp file to:", str(tmp_path))

# ================================================================
# Lettura robusta: distinguo Excel vs testo e parsing tollerante
# ================================================================
read_results = {}
warnings_list = []
detected_format = None
df = None

if is_xlsx_like:
    # proviamo Excel (openpyxl) prima
    try:
        df = pd.read_excel(BytesIO(raw_bytes), header=None, engine="openpyxl", dtype=str)
        detected_format = "excel_openpyxl_headerNone"
        read_results["pandas_openpyxl_headerNone"] = {"shape": df.shape, "head": df.head(5).astype(str).values.tolist()}
    except Exception as e:
        warnings_list.append(f"openpyxl read failed: {e}")
        read_results["pandas_openpyxl_headerNone_error"] = repr(e)
        df = None

    if df is None:
        # fallback openpyxl manual
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows = []
            maxcols = 0
            for row in ws.iter_rows(values_only=True):
                r = ["" if v is None else str(v) for v in row]
                rows.append(r)
                if len(r) > maxcols:
                    maxcols = len(r)
            norm = [r + [""] * (maxcols - len(r)) for r in rows]
            df = pd.DataFrame(norm).astype(str)
            detected_format = "excel_openpyxl_manual_rows"
            read_results["excel_openpyxl_manual_rows_head"] = df.head(5).astype(str).values.tolist()
        except Exception as e:
            warnings_list.append(f"openpyxl load_workbook failed: {e}")
            read_results["openpyxl_error"] = repr(e)
            df = None
else:
    # trattiamolo come testo/CSV
    try:
        try:
            text = raw_bytes.decode("utf-8")
            enc = "utf-8"
        except UnicodeDecodeError:
            try:
                text = raw_bytes.decode("utf-8-sig")
                enc = "utf-8-sig"
            except Exception:
                text = raw_bytes.decode("latin-1", errors="replace")
                enc = "latin-1"
        st.write("Decodifica effettuata con:", enc)

        # Primo tentativo: sniffer (ma non fidarsi completamente)
        try:
            sample = "\n".join(text.splitlines()[:20])
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            st.write("Sniffer CSV delimiter:", dialect.delimiter)
            try:
                df_csv = pd.read_csv(StringIO(text), sep=dialect.delimiter, header=None, dtype=str, engine="python", on_bad_lines="warn")
                read_results["pd_read_csv_sniff"] = {"shape": df_csv.shape, "head": df_csv.head(5).astype(str).values.tolist()}
                df = df_csv
                detected_format = f"csv_sniffer_sep_{repr(dialect.delimiter)}"
            except Exception as e:
                warnings_list.append(f"pd.read_csv with sniffer-delim failed: {e}")
                read_results["pd_read_csv_sniff_error_inner"] = repr(e)
        except Exception as e:
            read_results["pd_read_csv_sniff_error"] = repr(e)

        # Se il sniffer non ha prodotto df valido, utilizziamo il parser robusto
        if df is None:
            df, detected_format = try_parse_csv_text(text, warnings_list)

        # Se il parser produce una singola colonna con tab interni, forziamo split su tab
        if df is not None and df.shape[1] == 1:
            if df[0].str.contains("\t").any():
                df2 = df[0].str.split("\t", expand=True)
                if df2.shape[1] > 1:
                    df = df2.astype(str)
                    detected_format = (detected_format or "csv_singlecol") + "+forced_split_tab"
    except Exception as e:
        warnings_list.append(f"Impossibile interpretare come testo/CSV: {e}")
        read_results["csv_general_error"] = repr(e)
        df = None

# Normalizziamo NaN e assicuriamo stringhe
if df is not None:
    df = df.fillna("").astype(str)

# Output diagnostico
st.header("Risultati parsing preliminare")
st.write("detected_format:", detected_format)
if warnings_list:
    st.warning("Warnings durante il parsing:")
    for w in warnings_list:
        st.text(w)

st.write("read_results (dettagli):")
st.write(read_results)

if df is not None:
    st.write("✅ Anteprima (prime 20 righe, header NON interpretato):")
    st.dataframe(df.head(20), use_container_width=True)
    st.write(f"Dimensione DataFrame: {df.shape[0]} righe × {df.shape[1]} colonne")
    st.write("Prime 20 righe (visualizzazione testuale):")
    rows_preview = df.head(20).astype(str).values.tolist()
    for r in rows_preview:
        st.text(" | ".join(r))
else:
    st.error("Non è stato possibile ottenere un DataFrame dal file caricato con le strategie tentate.")

# Tentativo di individuare riga header candidata
if df is not None:
    hdr_idx = find_header_row_candidate(df)
    if hdr_idx is not None:
        st.info(f"Possibile header trovato alla riga {hdr_idx}. Valori header: {df.iloc[hdr_idx].tolist()}")
    else:
        st.info("Nessun header automatico trovato; usa il selettore per indicare quale riga usare.")

st.markdown("---")
st.info("Se le intestazioni sono in una riga interna, indica qui quale riga usare come header (es. 0 per prima riga, 5 per sesta), o premi il pulsante per procedere con l'elaborazione grezza.")

# CONTROLLO: input riga header da usare
if df is not None:
    header_row_input = st.number_input("Riga da usare come header (opzionale, -1 = nessuna)", min_value=-1, max_value=max(0, df.shape[0]-1), value=-1, step=1)
    use_header = None if header_row_input < 0 else int(header_row_input)
else:
    use_header = None

if st.button("Mostra anteprima interpretata con la riga di header scelta") and df is not None:
    try:
        if use_header is None:
            st.error("Nessuna riga di header scelta.")
        else:
            header_vals = list(df.iloc[use_header].astype(str).tolist())
            df_interp = df.copy()
            df_interp.columns = [f"col_{i}" for i in range(df.shape[1])]
            df_interp = df_interp.drop(index=use_header).reset_index(drop=True)
            df_interp.columns = header_vals
            st.write(f"Anteprima interpretata usando la riga {use_header} come header:")
            st.dataframe(df_interp.head(40), use_container_width=True)
    except Exception as e:
        st.error(f"Errore interpretazione header: {e}")

# Opzione: procedere con l'elaborazione normale usando il BytesIO originale o il tmp path
if st.button("Procedi con l'elaborazione (usando il file caricato)"):
    with st.spinner("Elaborazione in corso..."):
        try:
            codes_file = Path(__file__).parent / "codes.csv"
            codes_map = load_codes_map(codes_file)[0] if codes_file.exists() else {}
            # Proviamo prima a passare file-like (come fai ora)
            bio = BytesIO(raw_bytes)
            bio.seek(0)
            grouped_df, df_valid, inferred_month_str = process_workbook(
                bio,
                codes_map,
                infer_month=True,
                month_for_days=None,
                year_for_days=None
            )
            st.success("Elaborazione (file-like) completata.")
            st.subheader("Anteprima riepilogo (aggregato) - file-like")
            st.dataframe(grouped_df, use_container_width=True)
            st.subheader("Dati validi (anteprima) - file-like")
            st.dataframe(df_valid.head(200), use_container_width=True)
        except Exception as e:
            st.error(f"Errore durante l'elaborazione (file-like): {e}")

        # Se l'elaborazione file-like non ha prodotto dati utili oppure vuoi provare col path
        with st.spinner("Tentativo alternativo: passaggio del path al process_workbook..."):
            try:
                res2 = process_workbook(str(tmp_path), codes_map, infer_month=True, month_for_days=None, year_for_days=None)
                # res2 dovrebbe essere tuple (grouped_df, df_valid, inferred_month_str)
                if isinstance(res2, tuple) and len(res2) >= 2:
                    st.success("Elaborazione (path) completata.")
                    st.subheader("Anteprima riepilogo (aggregato) - path")
                    st.dataframe(res2[0], use_container_width=True)
                    st.subheader("Dati validi (anteprima) - path")
                    st.dataframe(res2[1].head(200), use_container_width=True)
                else:
                    st.warning("process_workbook(path) non ha restituito una struttura attesa.")
                    st.write("Raw result (repr):", repr(res2)[:2000])
            except Exception as e:
                st.error(f"Errore durante l'elaborazione (path): {e}")

st.success("DEBUG completato. Incolla qui i risultati principali (detected_format / warnings / shapes / error logs) per proseguire.")
